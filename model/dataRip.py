from zipfile import BadZipFile

import numpy as np
import pandas as pd
from binance import Client
from datetime import datetime as dt
from datetime import timedelta

# establish connection to Binance
api_key = "J0eCN1nkfjoCPVI3fHzBqU3yzscfgpoBYozm3sJpmunvhtxhKYUoh7IffisZ6yHk"
api_secret = "fqYiy16MWUAi61khQb6UGHH6M8XH6kFHPd6578aDZsGsHcDWBGnKf9tyOLH1EBEB"

client = Client(api_key, api_secret)


# append data to excel file
def append_df_to_excel(filename, df, sheet_name='Sheet', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    from openpyxl import load_workbook, Workbook

    import pandas as pd

    # ignore engine parameter if passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    try:
        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')
    except FileNotFoundError:
        wb = Workbook()
        wb.save(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, it will be created
        pass

    if startrow is None:
        startrow = 0

    # write out new sheet
    df.to_excel(writer,sheet_name, index=False, header=None, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

# get data from Binance
def scrape_hourly_candles(market="BTCUSDT", startingDate=dt(2018,8,7)):
    # set date to scrape from
    date = startingDate

    # set up all dates to scrape in an array
    dateList = []
    while date < dt.today():
        date += timedelta(days=1)
        dateList.append(dt.strftime(date, "%Y-%m-%d"))

    i = 0
    row = 1

    # scrape klines from binance in hourly intervals for each date in dateList
    while i < len(dateList) - 1:
        klines = client.get_historical_klines(market, client.KLINE_INTERVAL_1HOUR, start_str=dateList[i],
                                              end_str=dateList[i + 1])
        
        try:
            klines.pop(0)
        except:
            pass
        k = np.array(klines)
        times = []
        for j in range(0, 24):
            time = str(j) + ":00:00"
            if len(time) < 8:
                time = "0" + time
            times.append(dateList[i] + " " + time)

        if len(k) == 24:
            df = pd.DataFrame({"date": times,
                               "volume": k[:, 0],
                               "open": k[:, 1],
                               "high": k[:, 2],
                               "low": k[:, 3],
                               "close": k[:, 4]
                               })

            append_df_to_excel("./data/" + market + ".xlsx", df, startrow=row)
            row += 24
        i += 1
        print("saved " + str(row) + "rows of data")

if __name__ == "__main__":
    scrape_hourly_candles(market="ETHUSDT", startingDate=dt(2018,8,7))